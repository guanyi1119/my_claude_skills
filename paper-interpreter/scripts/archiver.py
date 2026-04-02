#!/usr/bin/env python3
"""
论文归档管理脚本
用于管理论文归档的Excel文件和文件操作
"""

import os
import sys
import re
from datetime import datetime
import argparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.cell import WriteOnlyCell
except ImportError:
    print("错误: 需要安装openpyxl库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


def sanitize_filename(filename):
    """
    清理文件名，替换非法字符

    Args:
        filename: 原始文件名

    Returns:
        清理后的文件名
    """
    # 替换非法字符为下划线
    invalid_chars = r'[<>:"/\\|?*]'
    sanitized = re.sub(invalid_chars, '_', filename)
    # 替换连续的多个下划线
    sanitized = re.sub(r'_+', '_', sanitized)
    # 去除首尾的空格和下划线
    sanitized = sanitized.strip(' _')
    return sanitized


def get_archive_paths(base_dir):
    """
    获取归档目录的路径

    Args:
        base_dir: 基础目录

    Returns:
        包含各种路径的字典
    """
    archive_dir = os.path.join(base_dir, 'archive')
    pdf_dir = os.path.join(archive_dir, 'pdf')
    md_dir = os.path.join(archive_dir, 'md')
    excel_path = os.path.join(archive_dir, 'paper_archive.xlsx')

    return {
        'archive_dir': archive_dir,
        'pdf_dir': pdf_dir,
        'md_dir': md_dir,
        'excel_path': excel_path
    }


def ensure_archive_directories(paths):
    """
    确保归档目录存在

    Args:
        paths: 路径字典
    """
    os.makedirs(paths['archive_dir'], exist_ok=True)
    os.makedirs(paths['pdf_dir'], exist_ok=True)
    os.makedirs(paths['md_dir'], exist_ok=True)


def create_or_load_excel(excel_path):
    """
    创建或加载Excel归档文件

    Args:
        excel_path: Excel文件路径

    Returns:
        Workbook对象
    """
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        return wb
    else:
        wb = Workbook()
        # 删除默认的Sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb


def get_or_create_sheet(wb, sheet_name):
    """
    获取或创建指定名称的Sheet

    Args:
        wb: Workbook对象
        sheet_name: Sheet名称

    Returns:
        Worksheet对象
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        # 添加表头
        headers = ['标题', '年份', '作者单位', '收录会议/杂志', '关键词', '归档原因', '归档日期', 'PDF路径', '解读报告路径']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        # 调整列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 40
        ws.column_dimensions['I'].width = 40
        return ws


def add_hyperlink(ws, row, col, display_text, file_path):
    """
    向单元格添加超链接

    Args:
        ws: Worksheet对象
        row: 行号
        col: 列号
        display_text: 显示文本
        file_path: 文件路径
    """
    # 使用绝对路径
    abs_path = os.path.abspath(file_path)
    # 转换为URL格式
    if sys.platform.startswith('win'):
        # Windows路径
        file_url = f'file:///{abs_path.replace("\\", "/")}'
    else:
        # Unix/Linux/Mac路径
        file_url = f'file://{abs_path}'

    cell = ws.cell(row=row, column=col)
    cell.value = display_text
    cell.hyperlink = file_url
    cell.style = 'Hyperlink'


def archive_paper(base_dir, paper_info, source_pdf_path, source_md_path, tag, archive_reason):
    """
    归档论文

    Args:
        base_dir: 基础目录
        paper_info: 论文信息字典
        paper_info = {
            'title': 论文标题,
            'year': 年份,
            'authors': 作者,
            'affiliation': 作者单位,
            'venue': 会议/杂志,
            'keywords': 关键词列表
        }
        source_pdf_path: 源PDF文件路径
        source_md_path: 源MD文件路径
        tag: 归档标签
        archive_reason: 归档原因

    Returns:
        成功返回True，失败返回False
    """
    try:
        # 获取路径
        paths = get_archive_paths(base_dir)
        ensure_archive_directories(paths)

        # 清理标题用于文件名
        sanitized_title = sanitize_filename(paper_info['title'])
        year = paper_info.get('year', '')

        # 生成目标文件名
        if year:
            pdf_filename = f"{sanitized_title}_{year}.pdf"
            md_filename = f"{sanitized_title}_{year}_解读.md"
        else:
            pdf_filename = f"{sanitized_title}.pdf"
            md_filename = f"{sanitized_title}_解读.md"

        target_pdf_path = os.path.join(paths['pdf_dir'], pdf_filename)
        target_md_path = os.path.join(paths['md_dir'], md_filename)

        # 复制文件
        import shutil
        if os.path.exists(source_pdf_path):
            shutil.copy2(source_pdf_path, target_pdf_path)
        if os.path.exists(source_md_path):
            shutil.copy2(source_md_path, target_md_path)

        # 更新Excel
        wb = create_or_load_excel(paths['excel_path'])
        ws = get_or_create_sheet(wb, tag)

        # 找到下一个空行
        next_row = ws.max_row + 1

        # 填写数据
        ws.cell(row=next_row, column=1, value=paper_info['title'])
        ws.cell(row=next_row, column=2, value=paper_info.get('year', ''))
        ws.cell(row=next_row, column=3, value=paper_info.get('affiliation', ''))
        ws.cell(row=next_row, column=4, value=paper_info.get('venue', ''))

        # 关键词（逗号分隔）
        keywords = paper_info.get('keywords', [])
        if isinstance(keywords, list):
            keywords_str = ', '.join(keywords)
        else:
            keywords_str = str(keywords)
        ws.cell(row=next_row, column=5, value=keywords_str)

        ws.cell(row=next_row, column=6, value=archive_reason)
        ws.cell(row=next_row, column=7, value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # 添加超链接
        if os.path.exists(target_pdf_path):
            add_hyperlink(ws, next_row, 8, '打开PDF', target_pdf_path)
        if os.path.exists(target_md_path):
            add_hyperlink(ws, next_row, 9, '打开解读', target_md_path)

        # 保存Excel
        wb.save(paths['excel_path'])

        print(f"论文已成功归档到标签: {tag}")
        print(f"Excel文件: {paths['excel_path']}")
        print(f"PDF文件: {target_pdf_path}")
        print(f"MD文件: {target_md_path}")

        return True

    except Exception as e:
        print(f"归档失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def delete_files(*file_paths):
    """
    删除指定文件

    Args:
        *file_paths: 要删除的文件路径列表
    """
    for file_path in file_paths:
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                print(f"已删除: {file_path}")
            except Exception as e:
                print(f"删除失败 {file_path}: {e}")


def main():
    """
    命令行接口（主要用于测试）
    """
    parser = argparse.ArgumentParser(description='论文归档管理工具')
    subparsers = parser.add_subparsers(dest='command', help='可用命令')

    # 测试创建Excel
    test_parser = subparsers.add_parser('test', help='测试归档功能')
    test_parser.add_argument('--base-dir', required=True, help='基础目录')
    test_parser.add_argument('--title', required=True, help='论文标题')
    test_parser.add_argument('--year', help='年份')
    test_parser.add_argument('--affiliation', help='作者单位')
    test_parser.add_argument('--venue', help='会议/杂志')
    test_parser.add_argument('--keywords', help='关键词（逗号分隔）')
    test_parser.add_argument('--tag', required=True, help='归档标签')
    test_parser.add_argument('--reason', required=True, help='归档原因')
    test_parser.add_argument('--pdf', help='PDF文件路径')
    test_parser.add_argument('--md', help='MD文件路径')

    args = parser.parse_args()

    if args.command == 'test':
        paper_info = {
            'title': args.title,
            'year': args.year,
            'affiliation': args.affiliation,
            'venue': args.venue,
            'keywords': args.keywords.split(',') if args.keywords else []
        }

        archive_paper(
            base_dir=args.base_dir,
            paper_info=paper_info,
            source_pdf_path=args.pdf,
            source_md_path=args.md,
            tag=args.tag,
            archive_reason=args.reason
        )


if __name__ == '__main__':
    main()
